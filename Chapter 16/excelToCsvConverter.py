#!python
# excelToCsvConverter.py Convierte archivos '.xlsx' a archivos '.csv'
# en lote desde un directorio.

import openpyxl, csv, re, os

# Crea el archivo de salida, si no existe.
os.makedirs('Csv_From_Excel', exist_ok=True)

# Crea la ruta de acceso y la lista de archivos.
path = r'C:\Users\tu_path\ExcelToCsv'
archivos = os.listdir(path)

# Itera sobre la lista de archivos y convierte de '.xlsx' a '.csv'.
for archivoExcel in archivos:
    # Carga el archivo Excel.
    wb = openpyxl.load_workbook(os.path.join(path, archivoExcel), read_only=True)
    # Itera sobre las hojas del archivo.
    for nombreHoja in wb.sheetnames:
        hoja = wb[nombreHoja]
        # Crea un aexpresión regex para extraer el nombre
        #  del archivo Excel.
        regEx = re.search(r'([^w]+?).xlsx?$', archivoExcel)
        # Crea el objeto 'csvWriter' con el nombre del archivo '.csv'. 
        nombreCsv = f'{regEx.group(1)}.csv'
        csvFileObj = open(os.path.join('Csv_From_Excel', nombreCsv), 'w', newline='')
        csvWriter = csv.writer(csvFileObj)

        # Itera por cada fila y columna copiando los valores en 'datos'.
        for filaNum in range(1, (hoja.max_row + 1)):
            datos = []
            for numColumn in range(1, (hoja.max_column + 1)):
                datos.append(hoja.cell(row=filaNum, column=numColumn).value)
            
        # Escribe los datos en el nuevo archivo y cierra el objeto.
            csvWriter.writerow(datos)
        csvFileObj.close()

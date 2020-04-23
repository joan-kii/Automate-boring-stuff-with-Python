#!python3
# blankRowInserter.py Inserta M filas en blanco a partir de la fila M.

import openpyxl, sys, re

# Obtiene el nombre del archivo, y la posición y el número de filas
# a insertar.
archivo = sys.argv[3]
N = int(sys.argv[1])
M = int(sys.argv[2])

# Carga el archivo y la hoja activa.
wb = openpyxl.load_workbook(archivo)
hoja = wb.active

# Itera desde N hasta N + M y sobre todas las columnas y escribe ''
# para crear las filas en blanco.
for numFila in range(N, N + M + 1):
    for numCol in range(1, hoja.max_column):
        if numFila >= N:
            hoja.cell(row=numFila, column=numCol).value = ''

# Crea una expresión regex para extraer el nombre del archivo original y
# copiarlo en el nombre del nuevo archivo más '_blank'.
regEx = re.search(r"([^w]+?).xlsx?$", archivo)
wb.save(f'{regEx.group(1)}_blank.xlsx')

# Cierra el archivo.
wb.close()
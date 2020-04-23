#!python3
# spreadsheetToTextFiles.py Convierte el texto de cada columna a un archivo txt.

import openpyxl, re

# Abre el archivo '.xlsx' y selecciona la hoja activa.
archivo = 'textToSpreadsheet.xlsx'
wb = openpyxl.load_workbook(archivo)
hoja = wb.active

# Itera por las celdas de la columna 1 de la hoja 'A' y si el valor de
# la celda es distinto de 'None', copia el valorde cada celda a la 
# lista 'texto1'.
texto1 = []
for filaNum1 in range(1, len(hoja['A']) + 1):
    if hoja.cell(row=filaNum1, column=1).value != None:
        texto1.append(hoja.cell(row=filaNum1, column=1).value)

# Itera por las celdas de la columna 2 de la hoja 'B' y si el valor de
# la celda es distinto de 'None', copia el valorde cada celda a la 
# lista 'texto2'.
texto2 = []
for filaNum2 in range(1, len(hoja['B']) + 1):
    if hoja.cell(row=filaNum2, column=2).value != None:
        texto2.append(hoja.cell(row=filaNum2, column=2).value)

# Itera por las celdas de la columna 3 de la hoja 'C' y si el valor de
# la celda es distinto de 'None', copia el valorde cada celda a la 
# lista 'texto3'.
texto3 = []
for filaNum3 in range(1, len(hoja['C']) + 1):
    if hoja.cell(row=filaNum3, column=3).value != None:
        texto3.append(hoja.cell(row=filaNum3, column=3).value)

# Crea una expresión regex para extraer el nombre del archivo original y
# copiarlo en el nombre del nuevo archivo más '_column_letra'.
regEx = re.search(r"([^w]+?).xlsx?$", archivo)

# Crea los nuevos archivos y escribe el contenido correspondiente.
nuevoArchivo1 = open(regEx.group(1) + '_column_A.txt', 'a')
nuevoArchivo1.write(''.join(texto1))

nuevoArchivo2 = open(regEx.group(1) + '_column_B.txt', 'a')
nuevoArchivo2.write(''.join(texto2))

nuevoArchivo3 = open(regEx.group(1) + '_column_C.txt', 'a')
nuevoArchivo3.write(''.join(texto3))

# Cierra los archivos.
nuevoArchivo1.close()
nuevoArchivo2.close()
nuevoArchivo3.close()
wb.close()
#!python3
# textFilesToSpreadsheet.py Convierte archivos de texto a spreadsheet
# separando los textos en columnas y las frases en celdas.

import openpyxl

# Crea el objeto 'Workbook' con 'openpyxl'.
wb = openpyxl.Workbook()
# Selecciona la hoja activa.
hoja = wb.active

# Abre los archivos a copiar.
archivo1 = open('archivo_1.txt', 'r')
archivo2 = open('archivo_2.txt', 'r')
archivo3 = open('archivo_3.txt', 'r')

# Convierte el texto de los archivos en listas de cada línea.
f1Lineas = archivo1.readlines()
f2Lineas = archivo2.readlines()
f3Lineas = archivo3.readlines()

# Itera por cada celda y copia el valor de la lista correspondiente.
for filaF1 in range(1, len(f1Lineas) + 1):
    hoja.cell(row=filaF1, column=1).value = f1Lineas[filaF1 - 1]

for filaF2 in range(1, len(f2Lineas) + 1):
    hoja.cell(row=filaF2, column=2).value = f2Lineas[filaF2 - 1]

for filaF3 in range(1, len(f3Lineas) + 1):
    hoja.cell(row=filaF3, column=3).value = f3Lineas[filaF3 - 1]

# Guarda el nuevo archivo '.xlsx'.
wb.save('textToSpreadsheet.xlsx')

# Cierra los archivos.
archivo1.close()
archivo2.close()
archivo3.close()
wb.close()
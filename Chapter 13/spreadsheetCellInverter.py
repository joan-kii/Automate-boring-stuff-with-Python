#!python3
# spreadsheetCellInverter.py Invierte columnas por filas.

import openpyxl, re
from openpyxl.utils import get_column_letter

# Abre el archivo y crea una nueva hoja.
archivo = 'example.xlsx'
wb = openpyxl.load_workbook(archivo)
hoja = wb.active
nuevaHoja = wb.create_sheet(index=0, title='inverted')

# Itera por todas la filas y columnas de 'hoja' y añade los valores de
# las celdas a 'celdas' y añade 'celdas' a 'filas'.
filas = []
for numFila in range(1, hoja.max_row + 1):
    celdas = []
    for numCol in range(1, hoja.max_column + 1):
        celdas.append(hoja.cell(row=numFila, column=numCol).value)
    filas.append(celdas)

# Itera por las filas y columnas de 'hoja' y copia los valores
# almacenados en 'filas'.
for nuevaFila in range(1, hoja.max_column + 1):
    for nuevaCol in range(1, hoja.max_row + 1):
        nuevaHoja.cell(row=nuevaFila, column=nuevaCol).value = filas[nuevaCol - 1][nuevaFila - 1]

# Itera sobre las columnas de 'nuevaHoja' y cambia el ancho de la columna.
for cellWidth in range(1, nuevaHoja.max_column + 1):
    column = get_column_letter(cellWidth)
    nuevaHoja.column_dimensions[column].width = 20

# Crea una expresión regex para extraer el nombre del archivo original y
# copiarlo en el nombre del nuevo archivo más '_inverted'.
regEx = re.search(r'([^w]+?).xlsx?$', archivo)
wb.save(f'{regEx.group(1)}_inverted.xlsx')

# Cierra el archivo.
wb.close()
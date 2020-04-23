#!python3
# multiplicationTableMaker.py Obtiene un número N desde cmd y crea una matriz de N x N.

import openpyxl, sys
from openpyxl.styles import Font

# Obtiene el número de 'command line'.
numero = int(sys.argv[1])

# Crea el objeto 'Workbook' de 'openpyxl'.
wb = openpyxl.Workbook()

# Selecciona la hoja activa.
hoja = wb.active

# Esrablece el estilo de la fuente a negrita.
fontObj = openpyxl.styles.Font(bold=True)

# Itera sobre filas y columnas para ir de celda en celda.
for numFila in range(1, numero + 2):
    for numCol in range(1, numero + 2):

        # En la primera celda escribe el número.
        if numFila == 1 and numCol == 1:
            hoja.cell(row=numFila, column=numCol).value = numero 

        # En la primera fila escribe el número correspondiente
        # de la matriz en negrita.
        elif numFila == 1:
            hoja.cell(row=numFila, column=numCol).value = numCol - 1
            hoja.cell(row=numFila, column=numCol).font = fontObj

        # En la primera columna escribe el número correspondiente
        # de la matriz en negrita.
        elif numCol == 1:
            hoja.cell(row=numFila, column=numCol).value = numFila - 1
            hoja.cell(row=numFila, column=numCol).font = fontObj

        # En el resto de celdas de la matriz escribe el resultado de la
        # multiplicación del número de la fila por el número de la
        # columna.
        else:
            hoja.cell(row=numFila, column=numCol).value = (numFila - 1) * (numCol - 1)

# Guarda el archivo como '.xlsx'.
wb.save(f'multiplicationTable_{numero}.xlsx')

# Cierra el archivo.
wb.close()